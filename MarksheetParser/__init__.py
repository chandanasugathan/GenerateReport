import openpyxl
from pathlib import Path
from datetime import date


class MarksheetParser:
    # Load the Workbook by Name
    input_filename = ""

    def __init__(self, filename):
        self.input_filename = filename

    def parse_sheet(self):
        print("Parsing through Excel....")
        xlsx_file = Path(self.input_filename)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.worksheets[0]
        return sheet

    def generate_column_mapping(self):
        sheet = self.parse_sheet()
        print("Columns...")
        iteration = 0
        column_list= dict()
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                column_list[cell.value] = iteration
                iteration = iteration + 1
        print(column_list)
        return column_list

    # Get List of Valid Subjects
    def get_subject_list(self):
        print("Getting List of Subjects...")
        sheet = self.parse_sheet()
        subject_column_list = ["G", "H", "I", "J", "K", "L", "M", "N"]
        subject_list = []
        for i in subject_column_list:
            if sheet[i][1].value is not None:
                subject_tuple = (i, sheet[i][0].value)
                subject_list.append(subject_tuple)
        print("Subjects List...")
        print(subject_list)
        return subject_list

    #Generating Data into Custom Template for Report Generation.
    def generate_report_data(self):
        report = []
        sheet = self.parse_sheet()
        subject_list = self.get_subject_list()
        column_dict = self.generate_column_mapping()
        m_row = sheet.max_row
        print("Maximum Rows = {0}".format(m_row))
        print("Generating Report Data....")
        for student in sheet.iter_rows(min_row=2, max_row=m_row):
            template = dict()
            template["date"] = '{:%d-%b-%Y}'.format(date.today())
            if student[int(column_dict["Roll_no"])].value is not None:    template['roll'] = str(student[1].value)
            if student[int(column_dict["Name of the student"])].value is not None:    template['name'] = str(student[3].value)
            if student[int(column_dict["Term"])].value is not None:    template['term'] = str(student[22].value)
            if student[int(column_dict["Branch"])].value is not None:    template['branch'] = str(student[23].value)
            if student[int(column_dict["Semester"])].value is not None:    template['sem'] = str(student[24].value)
            if student[int(column_dict["Division"])].value is not None:    template['division'] = str(student[25].value)
            table_entries = []
            iter = 1
            col = 6
            for subject in subject_list:
                subject_dict = dict()
                subject_dict['sr_no'] = str(iter)
                subject_dict['subject'] = str(subject[1])
                if student[col].value is not None:    subject_dict['mark'] = str(student[col].value)
                if student[col + 8].value is not None:    subject_dict['attendance'] = '{:.2f}'.format(
                    student[col + 8].value)
                col = col + 1
                iter = iter + 1
                table_entries.append(subject_dict)
            template['sr_no'] = table_entries
            report.append(template)
        print(report)
        return report
